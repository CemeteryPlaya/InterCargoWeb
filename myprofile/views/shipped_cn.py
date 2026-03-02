from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import HttpResponseForbidden
from django.contrib import messages
from django.core.exceptions import ValidationError
from datetime import datetime
from collections import defaultdict
from myprofile.models import TrackCode
from register.models import UserProfile
from myprofile.views.utils import is_staff as _is_staff, send_grouped_notifications, add_bulk_result_messages


def _parse_xlsx(file):
    """Парсит xlsx файл: столбец A, начиная с A3, до пустой ячейки или '条码数:'."""
    import openpyxl
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    codes = []
    for row_num in range(3, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=1)
        value = cell.value
        if value is None:
            break
        value = str(value).strip()
        if not value:
            break
        if '条码数:' in value:
            break
        codes.append(value)
    wb.close()
    return codes


@login_required
def shipped_cn_view(request):
    if not _is_staff(request.user):
        return HttpResponseForbidden("У вас нет доступа к этой странице.")

    if request.method != 'POST':
        return render(request, "shipped_cn.html")

    update_date_str = request.POST.get('update_date')
    track_codes_raw = request.POST.get('track_codes', '').strip()
    xlsx_file = request.FILES.get('xlsx_file')

    if not update_date_str:
        messages.error(request, "Укажите дату обновления.")
        return redirect('shipped_cn')

    try:
        update_date = datetime.strptime(update_date_str, "%Y-%m-%d").date()
    except ValueError:
        messages.error(request, "Неверный формат даты.")
        return redirect('shipped_cn')

    # Собираем трек-коды из textarea
    track_codes = [line.strip() for line in track_codes_raw.splitlines() if line.strip()]

    # Добавляем трек-коды из xlsx
    if xlsx_file:
        try:
            xlsx_codes = _parse_xlsx(xlsx_file)
            track_codes.extend(xlsx_codes)
            if xlsx_codes:
                messages.info(request, f"Из файла загружено {len(xlsx_codes)} трек-кодов.")
        except Exception as e:
            messages.error(request, f"Ошибка при чтении xlsx файла: {e}")
            return redirect('shipped_cn')

    # Убираем дубликаты, сохраняя порядок
    seen = set()
    unique_codes = []
    for code in track_codes:
        if code not in seen:
            seen.add(code)
            unique_codes.append(code)
    track_codes = unique_codes

    if not track_codes:
        messages.error(request, "Введите трек-коды или загрузите xlsx файл.")
        return redirect('shipped_cn')

    status = 'shipped_cn'
    updated = 0
    created = 0
    skipped = 0
    errors = 0
    notif_counts = defaultdict(int)

    for code in track_codes:
        try:
            track = TrackCode.objects.get(track_code=code)
            old_status = track.status

            # Не откатываем если статус уже дальше
            if TrackCode.STATUS_ORDER.get(old_status, 0) >= TrackCode.STATUS_ORDER.get(status, 0):
                skipped += 1
                continue

            track.status = status
            track.update_date = update_date

            try:
                track.save()
                updated += 1
                if track.owner:
                    notif_counts[track.owner] += 1
            except ValidationError as e:
                messages.error(request, f"Ошибка при обновлении {code}: {e}")
                errors += 1
                continue

        except TrackCode.DoesNotExist:
            # Создаём трек-код без владельца
            try:
                TrackCode.objects.create(
                    track_code=code,
                    status=status,
                    update_date=update_date,
                    owner=None,
                    weight=None
                )
                created += 1
            except ValidationError as e:
                messages.error(request, f"Не удалось создать трек-код {code}: {e}")
                errors += 1

    send_grouped_notifications(notif_counts, status)
    add_bulk_result_messages(request, updated=updated, created=created,
                              skipped=skipped, errors=errors)

    return redirect('shipped_cn')
